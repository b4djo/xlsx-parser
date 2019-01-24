from openpyxl import load_workbook

wb = load_workbook('')

sheet = wb.get_sheet_by_name('Лист1')

import MySQLdb
db = MySQLdb.connect(host="localhost",
                     user="root",
                     passwd="root",
                     db="test",
                     charset='utf8',
                     use_unicode=True)
cur = db.cursor()
maxCountRow = sheet.max_row

for i in range(2, maxCountRow):
    value = sheet['C' + str(i)].value
    sql = """INSERT INTO table1 (name)
            VALUES ('%(value)s')
            """ % {"value": value}
    cur.execute(sql)

db.commit()
db.close()